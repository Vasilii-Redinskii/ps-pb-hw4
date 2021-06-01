#Функция словаря с ранжированием по максимальным значениям
def get_max_val (excel_data, item_name):
    
    # Преобразуем переменную excel_data в список словарей с помощью метода to_dict()
    # Результат передаем в переменную excel_data_dict
    data_dict = excel_data.to_dict(orient='records')

    item_dict = defaultdict(int) # Создаем словарь с заранее заданным типом значений
    
    for element in data_dict:
        # Добавляем элемент в словарь item_dict
        # element[item_name] - название поля вычисления
        # Если ключа с таким названием в item_dict нет, то будет значение 0,
        # таким образом мы просто увеличим его на 1  
        name_product_list = element[item_name].split(ITEM_SPLITTER)
        for name_product in name_product_list:
            item_dict[name_product] += 1
        #item_dict[element[item_name]] += 1
        
    item_counter = Counter(item_dict)
    return item_counter

# Импортируем библиотеку pandas
import pandas
from openpyxl import load_workbook
from collections import defaultdict, Counter
from datetime import datetime

GENDER_MALE = 'м'
GENDER_FEMALE = 'ж'
KEY_GENDER = 'Пол'
KEY_ITEM = 'Купленные товары'
KEY_DATE = 'Дата посещения'
KEY_BROWSER = 'Браузер'
ITEM_SPLITTER = ','


# Читаем файл ексль и результат передаем в переменную excel_data
# Переменная excel_data имеет тип <class 'pandas.core.frame.DataFrame'>
excel_data_frame = pandas.read_excel('logs.xlsx', sheet_name='log')
# добавляем колонку Месяц в датафрейм
excel_data_frame['Month'] = excel_data_frame[KEY_DATE].dt.month
# создаем список уникальных месяцев
month_list = sorted(excel_data_frame.Month.unique())

# создаем таблицу для мужчин
m_excel_data = excel_data_frame[excel_data_frame[KEY_GENDER]==GENDER_MALE]
# создаем таблицу для женщин
f_excel_data = excel_data_frame[excel_data_frame[KEY_GENDER]==GENDER_FEMALE]
# создаем таблицу проверрочную таблицу по месяцам
feb_excel_data = excel_data_frame[excel_data_frame["Month"]==month_list[1]]

# Получаем список браузеров по популярности
browser_counter = get_max_val(excel_data_frame, KEY_BROWSER)

# Получаем список товаров по популярности
product_counter = get_max_val(excel_data_frame, KEY_ITEM)

# Групперуем посещения по браузерам и месяцам и считаем кол-во
browser_by_month=excel_data_frame.groupby([KEY_BROWSER, pandas.Grouper(key='Month')])[KEY_BROWSER].count()

# Получаем список товаров по популярности февраль
feb_product_counter = get_max_val(feb_excel_data, KEY_ITEM)

#Проверочная печать
print(browser_by_month.get(key = 'Яндекс: мобильное приложение'))
print(feb_product_counter.most_common(7))
print(product_counter.most_common(7))
print(browser_by_month.get(key = browser_counter.most_common(7)[0][0]).get(key = month_list[0]))


#Открытие ексель файла для записи
filename ='report.xlsx'
wb = load_workbook(filename)
sheet = wb['Лист1']
#Запись данных в ексель файл
for i in range (7):
    sheet.cell(row=i+5, column=1).value = browser_counter.most_common(7)[i][0]
    sheet.cell(row=i+19, column=1).value = product_counter.most_common(7)[i][0]
    sheet.cell(row=i+5, column=15).value = browser_counter.most_common(7)[i][1]
    for j in range(len(month_list)):
        sheet.cell(row=i+5, column=j+3).value = browser_by_month.get(key = browser_counter.most_common(7)[i][0]).get(key = month_list[j])
        

wb.save(filename)

